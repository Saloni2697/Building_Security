from tkinter import *
from tkinter import messagebox
from face_lib import *
from openpyxl import load_workbook
from facedetection import *

def login(m4,emailid,pswd):
    print("Inside")
    eid=emailid.get()
    psd=pswd.get()
    if (len(eid)==0 or len(psd)==0):
        messagebox.showwarning("Empty Fields","All fields are mendatory..!!")

    else:
    # Reading email and psd from excel
        rbook=load_workbook("buildingsys.xlsx")
        rsheet=rbook.get_sheet_by_name('Sheet1')
        row=rsheet.max_row
        col=rsheet.max_column
        print("ROWS and COLS:= ",row,col)
        elist=[]
        plist=[]
        for i in range(2,row+1):
          print("i",i)
          v=rsheet.cell(row=i,column=3)
          v1=rsheet.cell(row=i,column=4)
          val=v.value
          val2=v1.value
          elist.append(val)
          plist.append(val2)
        print("neww",eid,psd)
        print('Your list',elist)
        print('Your plist',plist)


        #checking for authorization
        if eid in elist:
            for i in range(len(elist)):
                if psd in plist:
                    if psd==plist[i]:
                        print("ITS A MATCH..!!!")
                        messagebox.showinfo("Save Your Image","Press 'S' to save your image..!!")
                        m4.destroy()
                        fd(0)
                        print('Face detected..!!!')
                        face_recog()
                        
                else:
                    messagebox.showerror("Unauthorized","Please Enter correct Password..!!")



        else:
            messagebox.showerror("Unauthorized","Please Enter correct Mail Id..!!")

            

    
def sign_in():
    print("Heyyaaaaaa")

    m4=Tk()
    m4.geometry('600x300')
    m4.title('Sign_IN')
    
    emailid = StringVar()
    pswd = StringVar()

    Label(m4,text='Sign In Page',font='Helveticca 22 bold italic',width='28',height='2').grid(row=0,column=1,columnspan=3)
    Label(m4,text=" ",height=1,width=6).grid(row=1,column=0)
    
    Label(m4,text="Email Id*",font='ComicSansMS 10 bold',height='2').grid(row=2,column=1)
    Label(m4,text=" ",height=2,width=6).grid(row=2,column=2)
    Entry(m4,textvariable=emailid).grid(row=2,column=3)


    Label(m4,text="Password*",font='Arial 10 bold',height=2).grid(row=3,column=1)
    Entry(m4,textvariable=pswd,show="*").grid(row=3,column=3)

    Button(m4,text="Get Yourself Clicked",font='Arial 14 bold',height=2,command=lambda:login(m4,emailid,pswd),width=16).grid(row=11,column=1,columnspan=3)
    Label(m4,text=" ",height=2,width=6).grid(row=13,column=2)


#sign_in()
