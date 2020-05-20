import face_recognition
import cv2
from openpyxl import load_workbook
from python_arduino import *
from mail_attachment import *

def face_recog():


  def xlread():
    rbook=load_workbook("buildingsys.xlsx")
    rsheet=rbook.get_sheet_by_name('Sheet1')
    row=rsheet.max_row
    col=rsheet.max_column
    print("ROWS and COLS:= ",row,col)
    a=[]
    caa=[]
    for i in range(2,row+1):
      print("58in55")
      print("i",i)
      v=rsheet.cell(row=i,column=1)
      ca=rsheet.cell(row=i,column=6)
      val=v.value
      val2=ca.value
      a.append(val)
      caa.append(val2)

    return a,caa

  names,imagess=xlread()
  print(names[0])
  print(imagess)
  my_image = []
  my_face_encoding = []
  for i in range(0,len(names)):
    print('12345')
    my_image.append((r"Images/"+imagess[i]))
    load_image=face_recognition.load_image_file(my_image[i])
    my_face_encoding.append(face_recognition.face_encodings(load_image)[0])

  print((my_image))
  
  #video_capture = cv2.VideoCapture(0)
  print("Heyy")




  # Create arrays of known face encodings and their names
  known_face_encodings=my_face_encoding
  print(my_face_encoding,len(my_face_encoding))
  known_face_names = names
  print(known_face_names)

  # Initialize some variables
  face_locations = []
  face_encodings = []
  face_names = []
  process_this_frame = True

  while True:
    #try:
      # Grab a single frame of video
      frame = cv2.imread(r"C:\\Users\\Saloni\\Desktop\\Building_Security\\0.jpg")
      cv2.imshow("img",frame)
      # Resize frame of video to 1/4 size for faster face recognition processing
      #small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

      # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
      rgb_small_frame = frame[:, :, ::-1]

      # Only process every other frame of video to save time
      if process_this_frame:
          # Find all the faces and face encodings in the current frame of video
          face_locations = face_recognition.face_locations(rgb_small_frame)
          face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

          face_names = []
          for face_encoding in face_encodings:
              # See if the face is a match for the known face(s)
              matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
              name = "Unknown"

              # If a match was found in known_face_encodings, just use the first one.
              if True in matches:
                  first_match_index = matches.index(True)
                  name = known_face_names[first_match_index]
                  print("Matched",name)
                  py_ino()
                  #break
              else:
                    print("Unknown")
                    email1()
                    #break
                
                  

              face_names.append(name)

      process_this_frame = not process_this_frame


      # Display the results
      for (top, right, bottom, left), name in zip(face_locations, face_names):
          # Scale back up face locations since the frame we detected in was scaled to 1/4 size
          #top *= 4
          #right *= 4
          #bottom *= 4
          #left *= 4

          # Draw a box around the face
          cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

          # Draw a label with a name below the face
          cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
          font = cv2.FONT_HERSHEY_DUPLEX
          cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

      # Display the resulting image
      cv2.imshow('Video', frame)

      # Hit 'q' on the keyboard to quit!
      if cv2.waitKey(1) & 0xFF == ord('q'):
          break
    
      if name != "Unknown":
        print("Matched",name)
        #py_ino()
        break
      else:
        print("Unknown")
        #email1()
        break
  
    #except:
      #pass


  # Release handle to the webcam
  cv2.destroyAllWindows()

#face_recog()
