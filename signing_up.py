from tkinter import *
from facedetection import *
#import Upgraded_GUI_prog 
from tkinter import messagebox
from openpyxl import load_workbook
import smtplib,random


fn,ln,eid,psd,repsd,ph,count="","","","","",0,0


def click_pic(m1,sender,msg,r):
    global count
    count=count+1
    global fname,lname,emailid,pswd,repswd,phone
    global fn,ln,eid,psd,repsd,ph
    try:
        
        print("Inside")
        fn=fname.get()
        ln=lname.get()
        eid=emailid.get()
        psd=pswd.get()
        repsd=repswd.get()
        ph=phone.get()
        print(len(fn),len(ln),len(eid),len(psd),len(repsd),ph)
        if (len(fn)==0 or len(ln)==0 or len(eid)==0 or len(psd)==0 or len(repsd) == 0 or ph==0):
            messagebox.showwarning("Empty Fields","All fields are mendatory..!!")
            count-=1
            m1.destroy()
            sign_up(sender,msg,r)
            print('stage 1')
        elif(psd!=repsd):
            count-=1
            messagebox.showwarning("Passwords Mismatched","Please ensure your passwords matches..!!")
        else:
            #count-=1
            messagebox.showinfo("Save Your Image","Press 'S' to save your image..!!")
            fd(ph)
    except TclError:
        count-=1
        messagebox.showwarning("Phone Number","Please enter correct phone number..!!")


def submit(m1,eid,sender,msg,r,fn,ln,psd,ph):
    if count==0:
        print("Enter values first")
        messagebox.showwarning("Enter Data First","Please enter your data first and get clicked..!!")
    else:

        
        def save_data():
            otp=e1.get()
            if otp==r:            
                rbook=load_workbook("buildingsys.xlsx")
                rsheet=rbook.get_sheet_by_name('Sheet1')
                row=rsheet.max_row
                col=rsheet.max_column
                for i in range(row+1,row+2):
                    for j in range(col):
                        if j==0:
                            rsheet['A'+str(i)]=fn
                        elif j==1:
                            rsheet['B'+str(i)]=ln
                        elif j==2:
                            rsheet['C'+str(i)]=eid
                        elif j==3:
                            rsheet['D'+str(i)]=psd
                        elif j==4:
                            rsheet['E'+str(i)]=ph
                        elif j==5:
                            rsheet['F'+str(i)]=(str(ph)+".jpg")
                rbook.save("buildingsys.xlsx")
                print("Done..!!!")
                messagebox.showinfo("Status","Values Updated successfully")
                m2.destroy()
                
            else:
                print("Wrong OTP")
                messagebox.showerror("Wrong OTP","Your Entered OTP is WRONG..!!!")
                m2.destroy()
                Page1()

        
        sender = 'saloni26597@gmail.com'
        password = 'Saloni@123'
        s=smtplib.SMTP('smtp.gmail.com',587)
        s.starttls()
        s.login(sender,password)
        s.sendmail(sender,eid,msg+str(r))
        print("Mail Send")
        s.quit()
        print("Mail Logout")
        print("Random Value is:-",r)

        m1.destroy()
        m2=Tk()
        m2.geometry('400x200')
        m2.title('Enter OTP')

        global e1
        e1=IntVar()
        
        Label(m2, text='',height=2,width=6).grid(row=1,column=0)
        Label(m2, text='Please enter your OTP',font='Helveticca 22 bold italic').grid(row=0,column=1)
        Label(m2, text='',height=2).grid(row=1,column=1)
        Entry(m2,textvariable=e1).grid(row=3,column=1)
        Label(m2, text='',height=2).grid(row=4,column=1)
        Button(m2,text="Save Data",font='Helveticca 16 bold italic',command=save_data).grid(row=5,column=1)




def sign_up(sender,msg,r):
    global fname,lname,emailid,pswd,repswd,phone
    global fn,ln,eid,psd,repsd,ph
    
    print("Heyy")
    m1=Tk()
    m1.geometry('650x550')
    m1.title('Sign_Up')

    fname = StringVar()
    lname = StringVar()
    emailid = StringVar()
    pswd = StringVar()
    repswd = StringVar()
    phone = IntVar()


    Label(m1,text='Sign Up Page',font='Helveticca 22 bold italic',width='28',height='2').grid(row=0,column=1,columnspan=3)
    Label(m1,text=" ",height=1,width=8).grid(row=1,column=0)
    
    Label(m1,text="First Name*",font='ComicSansMS 10 bold',height='2').grid(row=2,column=1)
    Label(m1,text=" ",height=2,width=6).grid(row=2,column=2)
    Entry(m1,textvariable=fname).grid(row=2,column=3)

    Label(m1,text="Last Name",font='Arial 10 bold',height=2).grid(row=3,column=1)
    Entry(m1,textvariable=lname).grid(row=3,column=3)

    Label(m1,text="Email Id",font='Arial 10 bold',height=2).grid(row=4,column=1)
    Entry(m1,textvariable=emailid).grid(row=4,column=3)

    Label(m1,text="Password",font='Arial 10 bold',height=2).grid(row=5,column=1)
    Entry(m1,textvariable=pswd,show="*").grid(row=5,column=3)

    Label(m1,text="Retype Password",font='Arial 10 bold',height=2).grid(row=6,column=1)
    Entry(m1,textvariable=repswd,show="*").grid(row=6,column=3)

    Label(m1,text="Phone_No",font='Arial 10 bold',height=2).grid(row=7,column=1)
    Entry(m1,textvariable=phone).grid(row=7,column=3)

    Label(m1,text=" ",height=2,width=8).grid(row=8,column=0)    
    Button(m1,text="Get Yourself Clicked",font='Arial 14 bold',height=2,command=lambda:click_pic(m1,sender,msg,r)).grid(row=9,column=1,columnspan=3)
    Label(m1,text=" ",height=1,width=8).grid(row=10,column=0)
    Button(m1,text="Submit",font='Arial 14 bold',height=2,command=lambda:submit(m1,eid,sender,msg,r,fn,ln,psd,ph),width=10).grid(row=11,column=1,columnspan=3)
    Label(m1,text=" ",height=2,width=6).grid(row=13,column=2)

    print("loop end")



sender = 'saloni26597@gmail.com'
subject = 'OTP for Building Security System' #Line that causes trouble
msg = 'Subject:{}\n\n Your OTP to open Gate1 is '.format(subject)
r= random.randint(10000,99999)

#sign_up(sender,msg,r)

